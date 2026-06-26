"""
api/reports.py - PDF and Excel Report Generation
"""

from fastapi import APIRouter, HTTPException, Depends
from fastapi.responses import StreamingResponse
from bson import ObjectId
from datetime import datetime
import io

from app.services.auth_service import get_current_user
from app.database import get_collection

router = APIRouter()


@router.get("/pdf/{job_id}")
async def download_pdf_report(job_id: str, current_user: dict = Depends(get_current_user)):
    """Generate and download a PDF screening report for a job."""
    try:
        from reportlab.lib.pagesizes import letter, A4
        from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
        from reportlab.lib.units import inch
        from reportlab.lib import colors
        from reportlab.platypus import SimpleDocTemplate, Paragraph, Spacer, Table, TableStyle
        from reportlab.lib.enums import TA_CENTER, TA_LEFT
    except ImportError:
        raise HTTPException(status_code=500, detail="ReportLab not installed. Run: pip install reportlab")

    jobs_col = get_collection("jobs")
    resumes_col = get_collection("resumes")

    try:
        job = await jobs_col.find_one({"_id": ObjectId(job_id)})
    except Exception:
        raise HTTPException(status_code=400, detail="Invalid job ID")

    if not job:
        raise HTTPException(status_code=404, detail="Job not found")

    candidates = []
    async for r in resumes_col.find({"job_id": job_id}).sort("ats_score", -1):
        candidates.append(r)

    # Build PDF in memory
    buffer = io.BytesIO()
    doc = SimpleDocTemplate(buffer, pagesize=A4, rightMargin=40, leftMargin=40,
                            topMargin=60, bottomMargin=40)

    styles = getSampleStyleSheet()
    title_style = ParagraphStyle('Title', parent=styles['Title'],
                                  fontSize=20, spaceAfter=6, textColor=colors.HexColor('#1e40af'))
    heading_style = ParagraphStyle('Heading', parent=styles['Heading2'],
                                    fontSize=13, spaceAfter=4, textColor=colors.HexColor('#374151'))
    normal_style = styles['Normal']

    content = []

    # Title
    content.append(Paragraph("AI Resume Screener", title_style))
    content.append(Paragraph(f"Screening Report: {job['title']}", heading_style))
    content.append(Paragraph(f"Generated: {datetime.utcnow().strftime('%B %d, %Y %H:%M UTC')}", normal_style))
    content.append(Paragraph(f"Total Candidates: {len(candidates)}", normal_style))
    content.append(Spacer(1, 0.3 * inch))

    # Summary Table
    content.append(Paragraph("Candidate Rankings", heading_style))
    content.append(Spacer(1, 0.1 * inch))

    table_data = [["Rank", "Candidate", "Email", "ATS Score", "Recommendation"]]

    for i, r in enumerate(candidates):
        parsed = r.get("parsed_data", {})
        rec = r.get("recommendation", "Pending")
        score = r.get("ats_score", 0) or 0
        table_data.append([
            str(i + 1),
            parsed.get("name", "Unknown") or "Unknown",
            parsed.get("email", "N/A") or "N/A",
            f"{score:.1f}%",
            rec
        ])

    col_widths = [0.5*inch, 1.8*inch, 2.2*inch, 1.1*inch, 1.2*inch]
    table = Table(table_data, colWidths=col_widths)
    table.setStyle(TableStyle([
        ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#1e40af')),
        ('TEXTCOLOR', (0, 0), (-1, 0), colors.white),
        ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
        ('FONTSIZE', (0, 0), (-1, 0), 10),
        ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
        ('ROWBACKGROUNDS', (0, 1), (-1, -1), [colors.white, colors.HexColor('#f1f5f9')]),
        ('GRID', (0, 0), (-1, -1), 0.5, colors.HexColor('#e2e8f0')),
        ('FONTSIZE', (0, 1), (-1, -1), 9),
        ('TOPPADDING', (0, 0), (-1, -1), 6),
        ('BOTTOMPADDING', (0, 0), (-1, -1), 6),
    ]))

    content.append(table)
    content.append(Spacer(1, 0.3 * inch))

    # Individual summaries
    content.append(Paragraph("Candidate Details", heading_style))
    for i, r in enumerate(candidates[:10]):  # Limit to top 10
        parsed = r.get("parsed_data", {})
        name = parsed.get("name", "Unknown") or "Unknown"
        score = r.get("ats_score", 0) or 0
        rec = r.get("recommendation", "Pending")

        content.append(Spacer(1, 0.15 * inch))
        content.append(Paragraph(f"#{i+1} {name} — {score:.1f}% ({rec})", heading_style))

        if r.get("summary"):
            content.append(Paragraph(r["summary"], normal_style))

        matched = r.get("skill_match", [])
        missing = r.get("missing_skills", [])
        if matched:
            content.append(Paragraph(f"✓ Matched Skills: {', '.join(matched[:8])}", normal_style))
        if missing:
            content.append(Paragraph(f"✗ Missing Skills: {', '.join(missing[:5])}", normal_style))

    doc.build(content)
    buffer.seek(0)

    filename = f"screening_report_{job['title'].replace(' ', '_')}_{datetime.utcnow().strftime('%Y%m%d')}.pdf"

    return StreamingResponse(
        buffer,
        media_type="application/pdf",
        headers={"Content-Disposition": f"attachment; filename={filename}"}
    )


@router.get("/excel/{job_id}")
async def download_excel_report(job_id: str, current_user: dict = Depends(get_current_user)):
    """Generate and download an Excel screening report."""
    try:
        import openpyxl
        from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
        from openpyxl.utils import get_column_letter
    except ImportError:
        raise HTTPException(status_code=500, detail="openpyxl not installed. Run: pip install openpyxl")

    jobs_col = get_collection("jobs")
    resumes_col = get_collection("resumes")

    try:
        job = await jobs_col.find_one({"_id": ObjectId(job_id)})
    except Exception:
        raise HTTPException(status_code=400, detail="Invalid job ID")

    if not job:
        raise HTTPException(status_code=404, detail="Job not found")

    candidates = []
    async for r in resumes_col.find({"job_id": job_id}).sort("ats_score", -1):
        candidates.append(r)

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Screening Results"

    # Styles
    header_fill = PatternFill("solid", fgColor="1e40af")
    header_font = Font(color="FFFFFF", bold=True, size=11)
    selected_fill = PatternFill("solid", fgColor="d1fae5")
    maybe_fill = PatternFill("solid", fgColor="fef3c7")
    rejected_fill = PatternFill("solid", fgColor="fee2e2")

    # Title row
    ws.merge_cells("A1:J1")
    ws["A1"] = f"Screening Report: {job['title']}"
    ws["A1"].font = Font(bold=True, size=14, color="1e40af")
    ws["A2"] = f"Generated: {datetime.utcnow().strftime('%Y-%m-%d %H:%M UTC')} | Total: {len(candidates)}"

    # Headers
    headers = ["Rank", "Name", "Email", "Phone", "Skills Count",
               "Experience (Yrs)", "ATS Score", "Matched Skills", "Missing Skills", "Recommendation"]
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=4, column=col, value=header)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center")

    # Data rows
    for i, r in enumerate(candidates):
        parsed = r.get("parsed_data", {})
        row = i + 5
        score = r.get("ats_score", 0) or 0
        rec = r.get("recommendation", "Pending")

        ws.cell(row=row, column=1, value=i + 1)
        ws.cell(row=row, column=2, value=parsed.get("name", "Unknown"))
        ws.cell(row=row, column=3, value=parsed.get("email", ""))
        ws.cell(row=row, column=4, value=parsed.get("phone", ""))
        ws.cell(row=row, column=5, value=len(parsed.get("skills", [])))
        ws.cell(row=row, column=6, value=parsed.get("total_experience_years", 0))
        ws.cell(row=row, column=7, value=round(score, 1))
        ws.cell(row=row, column=8, value=", ".join(r.get("skill_match", [])[:6]))
        ws.cell(row=row, column=9, value=", ".join(r.get("missing_skills", [])[:5]))
        ws.cell(row=row, column=10, value=rec)

        fill = selected_fill if rec == "Selected" else (maybe_fill if rec == "Maybe" else rejected_fill)
        for col in range(1, 11):
            ws.cell(row=row, column=col).fill = fill

    # Column widths
    widths = [6, 20, 28, 15, 12, 15, 12, 35, 30, 14]
    for col, width in enumerate(widths, 1):
        ws.column_dimensions[get_column_letter(col)].width = width

    buffer = io.BytesIO()
    wb.save(buffer)
    buffer.seek(0)

    filename = f"screening_{job['title'].replace(' ', '_')}_{datetime.utcnow().strftime('%Y%m%d')}.xlsx"

    return StreamingResponse(
        buffer,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f"attachment; filename={filename}"}
    )
